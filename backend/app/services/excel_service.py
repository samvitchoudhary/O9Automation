"""
Excel Export Service for generating test case Excel files
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.models import TestCase, TestStep


class ExcelService:
    """Service for exporting test cases to Excel format"""
    
    def __init__(self):
        self.border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_excel(self, test_cases: list) -> Workbook:
        """
        Generate Excel workbook matching the exact template format
        
        Args:
            test_cases: List of TestCase objects to export
            
        Returns:
            openpyxl Workbook object
        """
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create Cover Page
        self._create_cover_page(wb, test_cases)
        
        # Create Test Runs sheet
        self._create_test_runs_sheet(wb, test_cases)
        
        return wb
    
    def _create_cover_page(self, wb: Workbook, test_cases: list):
        """Create the cover page sheet"""
        ws = wb.create_sheet("Cover Page")
        
        # Add project title
        ws['A1'] = "O9 Test Automation Platform"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Add test run information
        ws['A3'] = "Test Run Information"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A5'] = "Date:"
        ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A6'] = "Total Test Cases:"
        ws['B6'] = len(test_cases)
        
        ws['A7'] = "Version:"
        ws['B7'] = "1.0"
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_test_runs_sheet(self, wb: Workbook, test_cases: list):
        """Create the Test Runs sheet with exact column format"""
        ws = wb.create_sheet("Test Runs_1")
        
        # Define column headers matching the template
        headers = [
            "DBID",
            "Directory",
            "Id",
            "Name",
            "Status",
            "Test Case Version",
            "Assigned To",
            "Executed Start",
            "Executed End",
            "Planned Start Date",
            "Planned End Date",
            "Defects",
            "Defect IDs",
            "Requirements",
            "Test Step #",
            "Test Step Description",
            "Test Step Expected Result",
            "Transaction Code",
            "Step Owner",
            "Test Step Actual Result",
            "Test Step Status"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = self.border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Write test case data
        row = 2
        for test_case in test_cases:
            # Generate test case ID (TR-XXXXX format)
            test_case_id = f"TR-{10000 + test_case.id}"
            
            # Status mapping
            status_map = {
                "draft": "Draft",
                "approved": "Approved",
                "in_progress": "In Progress",
                "completed": "Completed",
                "failed": "Failed"
            }
            
            # Sort steps by step_number
            sorted_steps = sorted(test_case.steps, key=lambda x: x.step_number)
            
            # Write first row with test case metadata
            ws.cell(row=row, column=1, value="")  # DBID
            ws.cell(row=row, column=2, value="")  # Directory
            ws.cell(row=row, column=3, value=test_case_id)  # Id
            ws.cell(row=row, column=4, value=test_case.name)  # Name
            ws.cell(row=row, column=5, value=status_map.get(test_case.status.value, test_case.status.value))  # Status
            ws.cell(row=row, column=6, value=test_case.test_case_version)  # Test Case Version
            ws.cell(row=row, column=7, value=test_case.assigned_to or "")  # Assigned To
            ws.cell(row=row, column=8, value="")  # Executed Start
            ws.cell(row=row, column=9, value="")  # Executed End
            ws.cell(row=row, column=10, value="")  # Planned Start Date
            ws.cell(row=row, column=11, value="")  # Planned End Date
            ws.cell(row=row, column=12, value="")  # Defects
            ws.cell(row=row, column=13, value="")  # Defect IDs
            ws.cell(row=row, column=14, value=test_case.requirements or "")  # Requirements
            ws.cell(row=row, column=15, value=sorted_steps[0].step_number if sorted_steps else "")  # Test Step #
            ws.cell(row=row, column=16, value=sorted_steps[0].description if sorted_steps else "")  # Test Step Description
            ws.cell(row=row, column=17, value=sorted_steps[0].expected_result if sorted_steps else "")  # Test Step Expected Result
            ws.cell(row=row, column=18, value=sorted_steps[0].transaction_code or "")  # Transaction Code
            ws.cell(row=row, column=19, value="")  # Step Owner
            ws.cell(row=row, column=20, value=sorted_steps[0].actual_result or "")  # Test Step Actual Result
            ws.cell(row=row, column=21, value=sorted_steps[0].status.value.replace("_", " ").title() if sorted_steps else "")  # Test Step Status
            
            # Apply borders to first row
            for col in range(1, 22):
                ws.cell(row=row, column=col).border = self.border_style
            
            row += 1
            
            # Write subsequent rows for remaining steps (only step information)
            for step in sorted_steps[1:]:
                ws.cell(row=row, column=1, value="")  # DBID
                ws.cell(row=row, column=2, value="")  # Directory
                ws.cell(row=row, column=3, value="")  # Id (empty for subsequent steps)
                ws.cell(row=row, column=4, value="")  # Name
                ws.cell(row=row, column=5, value="")  # Status
                ws.cell(row=row, column=6, value="")  # Test Case Version
                ws.cell(row=row, column=7, value="")  # Assigned To
                ws.cell(row=row, column=8, value="")  # Executed Start
                ws.cell(row=row, column=9, value="")  # Executed End
                ws.cell(row=row, column=10, value="")  # Planned Start Date
                ws.cell(row=row, column=11, value="")  # Planned End Date
                ws.cell(row=row, column=12, value="")  # Defects
                ws.cell(row=row, column=13, value="")  # Defect IDs
                ws.cell(row=row, column=14, value="")  # Requirements
                ws.cell(row=row, column=15, value=step.step_number)  # Test Step #
                ws.cell(row=row, column=16, value=step.description)  # Test Step Description
                ws.cell(row=row, column=17, value=step.expected_result)  # Test Step Expected Result
                ws.cell(row=row, column=18, value=step.transaction_code or "")  # Transaction Code
                ws.cell(row=row, column=19, value="")  # Step Owner
                ws.cell(row=row, column=20, value=step.actual_result or "")  # Test Step Actual Result
                ws.cell(row=row, column=21, value=step.status.value.replace("_", " ").title())  # Test Step Status
                
                # Apply borders
                for col in range(1, 22):
                    ws.cell(row=row, column=col).border = self.border_style
                
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 22):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15

